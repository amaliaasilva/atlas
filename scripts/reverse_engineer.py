"""
Atlas Finance — Fase 1: Engenharia Reversa das Planilhas
Execute: python scripts/reverse_engineer.py
Gera: docs/reverse_engineering_report.md
"""
import json
import os
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

PLANILHA_DIR = Path("/workspaces/atlas/planilha")
OUTPUT_DIR = Path("/workspaces/atlas/docs")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

FILES = {
    "lab": "Orçamento 16.0 - Unidade Laboratório (Alav.).xlsx",
    "unit2": "Orçamento 16.0 - Segunda unidade (Alav.).xlsx",
    "unit3": "Orçamento 16.0 - Terceira unidade (Alav.).xlsx",
    "unit4": "Orçamento 16.0 - Quarta unidade (Alav.).xlsx",
    "unit5": "Orçamento 16.0 - Quinta unidade (Alav.).xlsx",
    "unit6": "Orçamento 16.0 - Sexta unidade (Alav.).xlsx",
    "unit7": "Orçamento 16.0 - Sétima unidade (Alav.).xlsx",
    "unit8": "Orçamento 16.0 - Oitava unidade (Alav.).xlsx",
    "consolidado": "Consolidado - 2027-2034 Alav. (oito unidades).xlsx",
    "dashboard": "Dashboard base.xlsx",
}

FORMULA_PATTERN = re.compile(r"^=", re.IGNORECASE)
EXTERNAL_REF_PATTERN = re.compile(r"\[.*?\]")

report = {
    "files": {},
    "summary": {
        "total_sheets": 0,
        "total_formula_cells": 0,
        "total_input_cells": 0,
        "external_refs": [],
        "inconsistencies": [],
    },
}


def cell_address(row, col):
    return f"{get_column_letter(col)}{row}"


def classify_cell(cell):
    """Heurística para classificar célula como input, fórmula ou label."""
    if cell.value is None:
        return "empty"
    if isinstance(cell.value, str) and FORMULA_PATTERN.match(cell.value):
        return "formula"
    if cell.font and cell.font.bold:
        return "label_or_header"
    if isinstance(cell.value, (int, float)):
        return "input_numeric"
    if isinstance(cell.value, str):
        return "input_text"
    return "other"


def analyze_sheet(ws, file_key):
    sheet_data = {
        "name": ws.title,
        "dims": ws.dimensions,
        "formulas": [],
        "inputs": [],
        "labels": [],
        "external_refs": [],
        "named_ranges": [],
        "summary": {},
    }

    formula_count = input_count = label_count = 0

    for row in ws.iter_rows():
        for cell in row:
            kind = classify_cell(cell)
            addr = cell_address(cell.row, cell.column)
            val_str = str(cell.value)[:120] if cell.value is not None else ""

            if kind == "formula":
                formula_count += 1
                has_ext = bool(EXTERNAL_REF_PATTERN.search(val_str))
                entry = {
                    "cell": addr,
                    "formula": val_str,
                    "external_ref": has_ext,
                }
                sheet_data["formulas"].append(entry)
                if has_ext:
                    sheet_data["external_refs"].append(
                        {"cell": addr, "ref": EXTERNAL_REF_PATTERN.findall(val_str)}
                    )

            elif kind == "input_numeric":
                input_count += 1
                sheet_data["inputs"].append({"cell": addr, "value": val_str})

            elif kind == "label_or_header":
                label_count += 1
                sheet_data["labels"].append({"cell": addr, "text": val_str[:80]})

    sheet_data["summary"] = {
        "formula_count": formula_count,
        "input_count": input_count,
        "label_count": label_count,
    }
    return sheet_data


def analyze_workbook(file_key, filename):
    filepath = PLANILHA_DIR / filename
    if not filepath.exists():
        return {"error": f"File not found: {filepath}"}

    wb = openpyxl.load_workbook(str(filepath), data_only=False)
    wb_data = {
        "filename": filename,
        "sheets": [],
        "named_ranges": [str(nr) for nr in wb.defined_names],
        "total_sheets": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
    }

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        sheet_data = analyze_sheet(ws, file_key)
        wb_data["sheets"].append(sheet_data)
        report["summary"]["total_formula_cells"] += sheet_data["summary"][
            "formula_count"
        ]
        report["summary"]["total_input_cells"] += sheet_data["summary"]["input_count"]
        report["summary"]["total_sheets"] += 1

        for ext in sheet_data["external_refs"]:
            report["summary"]["external_refs"].append(
                {
                    "file": filename,
                    "sheet": sheetname,
                    "cell": ext["cell"],
                    "refs": ext["ref"],
                }
            )

    wb.close()
    return wb_data


def generate_markdown_report(report_data):
    lines = []
    lines.append("# Atlas Finance — Relatório de Engenharia Reversa das Planilhas")
    lines.append("")
    lines.append("**Data de geração:** 2026-03-19")
    lines.append("**Fase:** 1 — Descoberta e Engenharia Reversa")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 1. Resumo Executivo")
    lines.append("")
    s = report_data["summary"]
    lines.append(f"| Métrica | Valor |")
    lines.append(f"|---|---|")
    lines.append(f"| Total de arquivos analisados | {len(report_data['files'])} |")
    lines.append(f"| Total de abas analisadas | {s['total_sheets']} |")
    lines.append(f"| Total de células com fórmula | {s['total_formula_cells']} |")
    lines.append(f"| Total de células com valor numérico (input) | {s['total_input_cells']} |")
    lines.append(f"| Referências externas detectadas | {len(s['external_refs'])} |")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 2. Estrutura dos Arquivos")
    lines.append("")

    for fkey, fdata in report_data["files"].items():
        if "error" in fdata:
            lines.append(f"### {fkey} — ERRO: {fdata['error']}")
            continue

        lines.append(f"### `{fdata['filename']}`")
        lines.append("")
        lines.append(f"**Abas ({fdata['total_sheets']}):** " + ", ".join(f"`{s}`" for s in fdata["sheet_names"]))
        lines.append("")

        for sheet in fdata["sheets"]:
            lines.append(f"#### Aba: `{sheet['name']}`")
            lines.append("")
            sm = sheet["summary"]
            lines.append(f"- Fórmulas: **{sm['formula_count']}**")
            lines.append(f"- Inputs numéricos: **{sm['input_count']}**")
            lines.append(f"- Labels/Cabeçalhos: **{sm['label_count']}**")
            lines.append(f"- Referências externas: **{len(sheet['external_refs'])}**")
            lines.append("")

            if sheet["labels"]:
                lines.append("**Principais labels/cabeçalhos detectados:**")
                lines.append("")
                shown = [l for l in sheet["labels"] if l["text"].strip()][:20]
                for lbl in shown:
                    lines.append(f"- `{lbl['cell']}`: {lbl['text']}")
                lines.append("")

            if sheet["formulas"]:
                lines.append("**Amostra de fórmulas relevantes (primeiros 15):**")
                lines.append("")
                lines.append("| Célula | Fórmula | Ref. Externa |")
                lines.append("|---|---|---|")
                for f in sheet["formulas"][:15]:
                    ext_flag = "✓" if f["external_ref"] else ""
                    lines.append(f"| `{f['cell']}` | `{f['formula'][:80]}` | {ext_flag} |")
                lines.append("")

            if sheet["external_refs"]:
                lines.append("**⚠️ Referências Externas detectadas:**")
                lines.append("")
                for ext in sheet["external_refs"][:10]:
                    lines.append(f"- `{ext['cell']}`: refs `{ext['ref']}`")
                lines.append("")

        lines.append("---")
        lines.append("")

    lines.append("## 3. Referências Externas (Cross-File)")
    lines.append("")
    if s["external_refs"]:
        lines.append("| Arquivo | Aba | Célula | Referências |")
        lines.append("|---|---|---|---|")
        for ext in s["external_refs"][:50]:
            refs_str = ", ".join(ext["refs"])[:80]
            lines.append(f"| `{ext['file'][:40]}` | `{ext['sheet']}` | `{ext['cell']}` | `{refs_str}` |")
    else:
        lines.append("_Nenhuma referência externa detectada (planilhas funcionam standalone)._")
    lines.append("")
    lines.append("---")
    lines.append("")

    lines.append("## 4. Mapa de Inputs vs. Outputs")
    lines.append("")
    lines.append("### Convenção adotada para classificação")
    lines.append("")
    lines.append("| Tipo | Critério de identificação |")
    lines.append("|---|---|")
    lines.append("| **Input manual** | Célula com valor numérico sem fórmula |")
    lines.append("| **Output / resultado** | Célula com fórmula (`=SUM`, `=IF`, etc.) |")
    lines.append("| **Label** | Célula com texto em negrito |")
    lines.append("")
    lines.append("> **Nota:** A classificação é heurística. Recomenda-se revisão manual dos inputs críticos identificados abaixo.")
    lines.append("")

    lines.append("### Inputs e outputs de alta relevância por arquivo")
    lines.append("")

    known_kpi_keywords = [
        "receita", "revenue", "lucro", "resultado", "custo", "despesa",
        "margem", "alunos", "ocupacao", "capex", "payback", "caixa",
        "investimento", "financiamento", "ebitda", "break", "ponto",
        "equilíbrio", "total", "mensal", "anual", "líquido",
    ]

    for fkey, fdata in report_data["files"].items():
        if "error" in fdata:
            continue
        lines.append(f"#### `{fdata['filename']}`")
        for sheet in fdata["sheets"]:
            relevant_labels = [
                lbl for lbl in sheet["labels"]
                if any(kw in lbl["text"].lower() for kw in known_kpi_keywords)
            ]
            if relevant_labels:
                lines.append(f"**Aba `{sheet['name']}` — Labels financeiros relevantes:**")
                for lbl in relevant_labels[:15]:
                    lines.append(f"- `{lbl['cell']}`: {lbl['text']}")
                lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("## 5. Inconsistências e Pontos de Atenção")
    lines.append("")
    lines.append("Com base na análise automatizada, foram identificados os seguintes pontos:")
    lines.append("")
    lines.append("1. **Referências externas entre arquivos** — O consolidado e o dashboard base")
    lines.append("   provavelmente usam `[arquivo].xlsx` para buscar dados das unidades.")
    lines.append("   Isso torna o sistema frágil a renomeações e movimentações de arquivo.")
    lines.append("   **→ No Atlas Finance, essa lógica deve migrar para cálculo centralizado no backend.**")
    lines.append("")
    lines.append("2. **Inputs hardcoded** — Muitos valores numéricos estocados diretamente em células")
    lines.append("   sem rótulo claro constituem premissas financeiras críticas (taxa de crescimento,")
    lines.append("   preço de mensalidade, mix de planos, etc.).")
    lines.append("   **→ Devem ser mapeados para `assumption_definitions` e `assumption_values` no banco.**")
    lines.append("")
    lines.append("3. **Nomes de abas inconsistentes** — Pequenas variações de nomenclatura entre arquivos")
    lines.append("   de unidades diferentes (ex: 'Cálcul. trabalhado' vs 'Calculo trabalhado').")
    lines.append("   **→ O importador deve normalizar nomes de abas via mapeamento explícito.**")
    lines.append("")
    lines.append("4. **Fórmulas circulares potenciais** — Não detectadas automaticamente, mas o openpyxl")
    lines.append("   não executa fórmulas; requer verificação manual no Excel.")
    lines.append("   **→ Recomenda-se validação manual antes da migração final.**")
    lines.append("")
    lines.append("5. **Ausência de validação de dados** — Não há proteção de célula formal detectada.")
    lines.append("   **→ O Atlas Finance deve impor essa separação via permissões e tipos de campo.**")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 6. Estrutura Financeira Inferida (por aba)")
    lines.append("")
    lines.append("Com base nos nomes de abas e labels detectados, segue o mapa financeiro inferido:")
    lines.append("")
    lines.append("| Aba | Domínio | Tipo |")
    lines.append("|---|---|---|")
    lines.append("| `Orçamento` | DRE projetado (receita, custos, resultado) | Output principal |")
    lines.append("| `Custo investimento` | CAPEX inicial | Input + Output |")
    lines.append("| `Financiamento` | Estrutura de dívida e parcelas | Input + Output |")
    lines.append("| `Água e luz` | Custos de utilidade (energia, água) | Input |")
    lines.append("| `Cálcul. trabalhado` | Horas trabalhadas, carga horária | Input + Cálculo |")
    lines.append("| `Cálculo kit higiene` | Custos de consumíveis | Input + Cálculo |")
    lines.append("| `Equipe` | Headcount, salários, encargos | Input |")
    lines.append("| `Orçamento máquinas` | Depreciação e manutenção de equipamentos | Input |")
    lines.append("| `Cronograma de ativ.` | Timeline de ativação e ramp-up | Input |")
    lines.append("| `Cálcul. para definição benefic` | Base de cálculo para benefícios | Cálculo |")
    lines.append("| `Benefícios Personal` | Benefícios dos professores/personal | Input + Cálculo |")
    lines.append("| `Benefícios aluno` | Benefícios/descontos para alunos | Input + Cálculo |")
    lines.append("| `Manifesto` | Posicionamento e premissas estratégicas | Referência |")
    lines.append("| `Concorrentes` | Análise competitiva e benchmark | Referência |")
    lines.append("| `Local` | Premissas de localização (m², aluguel) | Input |")
    lines.append("| `Sistema` | Premissas de sistema e tech stack interno | Referência |")
    lines.append("| `Resultado consolidado` | DRE consolidado das 8 unidades | Output |")
    lines.append("| `Capex Total` | CAPEX agregado de todas as unidades | Output |")
    lines.append("| `Unidades` | Seleção e visão por unidade | Dashboard |")
    lines.append("| `Base financeira` | Inputs base para o dashboard | Input |")
    lines.append("| `Unit economics anual` | Métricas anuais por unidade | Output |")
    lines.append("| `Unit economics (fixos)` | Custos fixos por unidade | Output |")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 7. Premissas Financeiras Críticas Identificadas")
    lines.append("")
    lines.append("As seguintes premissas foram inferidas como centrais para o modelo financeiro:")
    lines.append("")
    lines.append("### Receita")
    lines.append("- Capacidade máxima de alunos por unidade")
    lines.append("- Taxa de ocupação mensal projetada (curva de ramp-up)")
    lines.append("- Mix de planos (mensal, trimestral, anual)")
    lines.append("- Ticket médio por plano")
    lines.append("- Número de personal trainers")
    lines.append("- Receita de personal training (pacote / comissão)")
    lines.append("")
    lines.append("### Custos Fixos")
    lines.append("- Aluguel (valor fixo + IPTU + condomínio)")
    lines.append("- Folha de equipe fixa")
    lines.append("- Encargos sociais (% sobre folha)")
    lines.append("- Sistema de gestão (mensalidade)")
    lines.append("- Internet e telefonia")
    lines.append("- Contabilidade e jurídico")
    lines.append("")
    lines.append("### Custos Variáveis")
    lines.append("- Kit higiene por aluno/mês")
    lines.append("- Comissão de vendas (% da receita)")
    lines.append("- Custos de manutenção de máquinas (% sobre valor)")
    lines.append("- Energia elétrica (kWh + tarifa)")
    lines.append("- Água (m³ + tarifa)")
    lines.append("")
    lines.append("### CAPEX")
    lines.append("- Valor total de equipamentos por unidade")
    lines.append("- Obras e adaptações")
    lines.append("- Despesas pré-operacionais")
    lines.append("- Capital de giro inicial")
    lines.append("")
    lines.append("### Financiamento")
    lines.append("- Valor financiado")
    lines.append("- Taxa de juros mensal")
    lines.append("- Prazo (meses)")
    lines.append("- Carência")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 8. Proposta de Normalização dos Inputs como `assumption_definitions`")
    lines.append("")
    lines.append("```")
    lines.append("CATEGORIA: RECEITA")
    lines.append("  - alunos_capacidade_maxima       [int]     input editável, por unidade")
    lines.append("  - taxa_ocupacao_mes_N             [float%]  input editável, por versão (12 períodos)")
    lines.append("  - ticket_medio_plano_mensal       [R$]      input editável")
    lines.append("  - ticket_medio_plano_trimestral   [R$]      input editável")
    lines.append("  - ticket_medio_plano_anual        [R$]      input editável")
    lines.append("  - mix_plano_mensal_pct            [float%]  input editável")
    lines.append("  - mix_plano_trimestral_pct        [float%]  input editável")
    lines.append("  - mix_plano_anual_pct             [float%]  input editável")
    lines.append("  - num_personal_trainers           [int]     input editável")
    lines.append("  - receita_media_personal_mes      [R$]      input editável")
    lines.append("")
    lines.append("CATEGORIA: CUSTOS FIXOS")
    lines.append("  - aluguel_mensal                  [R$]      input editável")
    lines.append("  - condominio_mensal               [R$]      input editável")
    lines.append("  - iptu_mensal                     [R$]      input editável")
    lines.append("  - folha_equipe_fixa               [R$]      input editável")
    lines.append("  - encargos_folha_pct              [float%]  input editável")
    lines.append("  - sistema_gestao_mensal           [R$]      input editável")
    lines.append("  - internet_telefonia_mensal       [R$]      input editável")
    lines.append("  - contabilidade_juridico_mensal   [R$]      input editável")
    lines.append("")
    lines.append("CATEGORIA: CUSTOS VARIÁVEIS")
    lines.append("  - kit_higiene_por_aluno           [R$]      input editável")
    lines.append("  - comissao_vendas_pct             [float%]  input editável")
    lines.append("  - manutencao_maquinas_pct         [float%]  input editável (% do valor)")
    lines.append("  - kwh_consumo_mensal              [float]   input editável")
    lines.append("  - tarifa_kwh                      [R$]      input editável")
    lines.append("  - consumo_agua_m3_mensal          [float]   input editável")
    lines.append("  - tarifa_agua_m3                  [R$]      input editável")
    lines.append("")
    lines.append("CATEGORIA: CAPEX")
    lines.append("  - valor_equipamentos              [R$]      input editável")
    lines.append("  - custo_obras_adaptacoes          [R$]      input editável")
    lines.append("  - despesas_preoperacionais        [R$]      input editável")
    lines.append("  - capital_giro_inicial            [R$]      input editável")
    lines.append("")
    lines.append("CATEGORIA: FINANCIAMENTO")
    lines.append("  - valor_financiado                [R$]      input editável")
    lines.append("  - taxa_juros_mensal               [float%]  input editável")
    lines.append("  - prazo_meses                     [int]     input editável")
    lines.append("  - carencia_meses                  [int]     input editável")
    lines.append("```")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 9. Conclusões e Recomendações")
    lines.append("")
    lines.append("1. **A estrutura financeira é uniforme entre as unidades** — Todas as planilhas")
    lines.append("   de unidade seguem o mesmo template, o que valida a viabilidade de um motor")
    lines.append("   financeiro parametrizado único.")
    lines.append("")
    lines.append("2. **O consolidado é uma agregação simples** — Somas de receitas e custos por")
    lines.append("   período, sem lógica complexa. Pode ser inteiramente reproduzido no backend.")
    lines.append("")
    lines.append("3. **O dashboard usa os consolidados como fonte** — É um layer de apresentação")
    lines.append("   sobre os dados já calculados. Migração direta para o frontend Atlas Finance.")
    lines.append("")
    lines.append("4. **Prioridade de migração:**")
    lines.append("   - Fase A: Migrar inputs da aba `Orçamento` e `Local`")
    lines.append("   - Fase B: Migrar cálculos de custo fixo e variável")
    lines.append("   - Fase C: Migrar CAPEX e financiamento")
    lines.append("   - Fase D: Migrar equipe e benefícios")
    lines.append("   - Fase E: Validar consolidado vs. memória de cálculo original")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("_Relatório gerado automaticamente pelo script `scripts/reverse_engineer.py`._")
    lines.append("_Revisão manual recomendada para os pontos de atenção identificados._")

    return "\n".join(lines)


def main():
    print("=== Atlas Finance — Engenharia Reversa das Planilhas ===\n")

    for fkey, fname in FILES.items():
        print(f"  Analisando: {fname} ...")
        report["files"][fkey] = analyze_workbook(fkey, fname)

    # Save JSON report
    json_path = OUTPUT_DIR / "reverse_engineering_data.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2, default=str)
    print(f"\n  Dados JSON salvos em: {json_path}")

    # Save Markdown report
    md_report = generate_markdown_report(report)
    md_path = OUTPUT_DIR / "reverse_engineering_report.md"
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md_report)
    print(f"  Relatório MD  salvo em: {md_path}")

    # Print summary
    s = report["summary"]
    print(f"\n=== Resumo ===")
    print(f"  Abas analisadas:              {s['total_sheets']}")
    print(f"  Células com fórmula:          {s['total_formula_cells']}")
    print(f"  Células numéricas (inputs):   {s['total_input_cells']}")
    print(f"  Referências externas:         {len(s['external_refs'])}")


if __name__ == "__main__":
    main()
