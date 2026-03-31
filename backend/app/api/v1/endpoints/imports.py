"""
Atlas Finance — Import Endpoint
Upload e processamento de arquivos Excel.
"""

import shutil
import uuid
import io
from pathlib import Path
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.config import settings
from app.api.v1.deps import get_current_user
from app.models.user import User
from app.importers.excel_importer import import_from_path

router = APIRouter()


@router.post("/upload/{unit_id}")
async def upload_excel(
    unit_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Faz upload de um arquivo Excel e inicia o pipeline de importação
    para a unidade especificada (usa a budget_version ativa da unidade).
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400, detail="Apenas arquivos Excel (.xlsx, .xls) são permitidos"
        )

    # Busca a budget_version ativa da unidade
    from app.models.budget_version import BudgetVersion

    budget_version = (
        db.query(BudgetVersion)
        .filter(BudgetVersion.unit_id == unit_id, BudgetVersion.is_active == True)
        .order_by(BudgetVersion.created_at.desc())
        .first()
    )
    if not budget_version:
        raise HTTPException(
            status_code=404,
            detail="Nenhuma budget version ativa encontrada para esta unidade",
        )

    # Sanitiza o nome do arquivo
    safe_name = f"{uuid.uuid4()}_{Path(file.filename).name}"
    upload_dir = Path(settings.UPLOAD_DIR)
    upload_dir.mkdir(parents=True, exist_ok=True)
    dest = upload_dir / safe_name

    try:
        with open(dest, "wb") as f:
            shutil.copyfileobj(file.file, f)
    finally:
        file.file.close()

    result = import_from_path(str(dest), budget_version.id, db, unit_id=unit_id)
    return result


@router.get("/jobs")
def list_import_jobs(
    unit_id: str | None = Query(None),
    budget_version_id: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.import_job import ImportJob

    q = db.query(ImportJob)
    if unit_id:
        q = q.filter(ImportJob.unit_id == unit_id)
    if budget_version_id:
        q = q.filter(ImportJob.budget_version_id == budget_version_id)
    return q.order_by(ImportJob.created_at.desc()).limit(50).all()


@router.get("/jobs/{job_id}")
def get_import_job(
    job_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.import_job import ImportJob

    job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job não encontrado")
    return job


@router.get("/template")
def download_template(
    current_user: User = Depends(get_current_user),
):
    """Gera e retorna um arquivo Excel (.xlsx) com o modelo de importação."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import date

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove sheet padrão

    HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    LABEL_FONT = Font(bold=True)
    MONTHS = [date(2026, m, 1) for m in range(1, 13)]

    SHEETS = {
        "Local": [
            "Aluguel",
            "Condomínio",
            "IPTU",
            "Área (m²)",
        ],
        "Equipe": [
            "Limpeza",
            "Recepção",
            "Marketing",
            "Comercial",
            "Gerente",
            "Educador Físico",
            "Pró-labore",
            "Encargos (%)",
            "Benefícios por funcionário",
        ],
        "Água e Luz": [
            "Consumo kWh mensal",
            "Tarifa kWh",
            "Consumo Água m3 mensal",
            "Tarifa Água m3",
            "Internet / Telefonia",
        ],
        "Custo Investimento": [
            "Equipamentos",
            "Custo Obras / Adaptações",
            "Despesas Pré-operacionais",
            "Capital de Giro Inicial",
            "Móveis e Fixtures",
            "Tecnologia Setup",
        ],
        "Financiamento": [
            "Valor Financiado",
            "Taxa Juros Mensal (%)",
            "Prazo (meses)",
            "Carência (meses)",
        ],
    }

    for sheet_name, labels in SHEETS.items():
        ws = wb.create_sheet(sheet_name)

        # Cabeçalho: col A = "Descrição", cols B.. = meses
        ws.cell(1, 1, "Descrição").font = HEADER_FONT
        ws.cell(1, 1).fill = HEADER_FILL
        ws.cell(1, 1).alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 32

        for col_idx, month in enumerate(MONTHS, start=2):
            cell = ws.cell(1, col_idx, month.strftime("%Y-%m"))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            col_letter = ws.cell(1, col_idx).column_letter
            ws.column_dimensions[col_letter].width = 14

        # Linhas de premissas
        for row_idx, label in enumerate(labels, start=2):
            cell = ws.cell(row_idx, 1, label)
            cell.font = LABEL_FONT
            # Deixa as células de valor formatadas como número
            for col_idx in range(2, len(MONTHS) + 2):
                ws.cell(row_idx, col_idx).number_format = "#,##0.00"

        # Nota de instrução
        note_row = len(labels) + 3
        ws.cell(note_row, 1, "⚠ Preencha os valores mensais nas colunas à direita de cada linha.")
        ws.cell(note_row, 1).font = Font(italic=True, color="888888")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_importacao_atlas.xlsx"},
    )
