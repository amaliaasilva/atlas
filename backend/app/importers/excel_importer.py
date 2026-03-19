"""
Atlas Finance — Excel Importer
Pipeline de importação das planilhas legadas para o banco relacional.

Fluxo:
  1. Recebe o upload do arquivo Excel
  2. Registra ImportJob
  3. Lê abas e células relevantes
  4. Mapeia para assumption_values
  5. Persiste os dados
  6. Atualiza status do ImportJob
"""
from __future__ import annotations

import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl

if TYPE_CHECKING:
    from sqlalchemy.orm import Session

# Mapeamento: (nome_aba, linha, coluna_label) → assumption_code
# Construído a partir da engenharia reversa das planilhas
# Formato: (sheet_name_normalized, row_index, label_keyword) → code

SHEET_MAP = {
    "local": {
        "aluguel": "aluguel_mensal",
        "condomínio": "condominio_mensal",
        "iptu": "iptu_mensal",
        "área": "area_m2",
        "m²": "area_m2",
    },
    "equipe": {
        "limpeza": "salario_limpeza",
        "recepção": "salario_recepcao",
        "recepcao": "salario_recepcao",
        "marketing": "salario_marketing",
        "comercial": "salario_comercial",
        "gerente": "salario_gerente",
        "educador": "salario_educador_fisico",
        "pró-labore": "pro_labore",
        "prolabore": "pro_labore",
        "encargos": "encargos_folha_pct",
        "benefícios": "beneficios_por_funcionario",
        "beneficios": "beneficios_por_funcionario",
    },
    "água e luz": {
        "kwh": "kwh_consumo_mensal",
        "tarifa": "tarifa_kwh",
        "água": "consumo_agua_m3_mensal",
        "agua": "consumo_agua_m3_mensal",
        "m3": "tarifa_agua_m3",
        "internet": "internet_telefonia_mensal",
        "telefone": "internet_telefonia_mensal",
    },
    "custo investimento": {
        "equipamento": "valor_equipamentos",
        "obra": "custo_obras_adaptacoes",
        "reforma": "custo_obras_adaptacoes",
        "pré-operacional": "despesas_preoperacionais",
        "preoperacional": "despesas_preoperacionais",
        "capital de giro": "capital_giro_inicial",
        "giro": "capital_giro_inicial",
        "móveis": "moveis_e_fixtures",
        "moveis": "moveis_e_fixtures",
        "tecnologia": "tecnologia_setup",
    },
    "financiamento": {
        "valor financiado": "valor_financiado",
        "financiado": "valor_financiado",
        "taxa": "taxa_juros_mensal",
        "prazo": "prazo_meses",
        "carência": "carencia_meses",
        "carencia": "carencia_meses",
    },
}


def _normalize(text: str) -> str:
    """Normaliza string para matching."""
    return re.sub(r"\s+", " ", str(text).lower().strip())


def _find_assumption_code(sheet_name: str, label: str) -> str | None:
    """Tenta encontrar o código da premissa pelo nome da aba e o label da linha."""
    norm_sheet = _normalize(sheet_name)
    norm_label = _normalize(label)

    for sheet_key, mappings in SHEET_MAP.items():
        if sheet_key in norm_sheet:
            for keyword, code in mappings.items():
                if keyword in norm_label:
                    return code
    return None


def _extract_period_from_header(header_value) -> str | None:
    """Tenta extrair YYYY-MM de um cabeçalho de coluna."""
    if header_value is None:
        return None
    s = str(header_value)
    # Tenta datetime
    if hasattr(header_value, "year"):
        return f"{header_value.year:04d}-{header_value.month:02d}"
    # Tenta string ISO
    m = re.match(r"(\d{4})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return None


class ExcelImporter:
    """
    Importa um arquivo Excel de orçamento de unidade para o banco de dados.
    """

    def __init__(self, file_path: str, budget_version_id: str, db: "Session"):
        self.file_path = file_path
        self.budget_version_id = budget_version_id
        self.db = db
        self.errors: list[str] = []
        self.imported_count = 0

    def run(self) -> dict:
        """Executa o pipeline de importação."""
        from app.models.import_job import ImportJob, ImportStatus
        from app.models.assumption import AssumptionValue, AssumptionDefinition

        # Registra o job
        job_id = str(uuid.uuid4())
        job = ImportJob(
            id=job_id,
            file_name=Path(self.file_path).name,
            file_path=self.file_path,
            import_type="unit_budget",
            status=ImportStatus.processing,
            budget_version_id=self.budget_version_id,
            started_at=datetime.now(timezone.utc),
        )
        self.db.add(job)
        self.db.commit()

        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            discovered = {}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_results = self._process_sheet(ws, sheet_name)
                discovered.update(sheet_results)

            # Valida premissas contra as definições no banco
            defs = (
                self.db.query(AssumptionDefinition)
                .all()
            )
            code_to_def = {d.code: d for d in defs}

            values_to_insert = []
            for (code, period), value in discovered.items():
                defn = code_to_def.get(code)
                if not defn:
                    continue
                values_to_insert.append(
                    AssumptionValue(
                        budget_version_id=self.budget_version_id,
                        assumption_definition_id=defn.id,
                        period_date=period,
                        value_numeric=float(value) if value is not None else None,
                        source_type="imported",
                        source_reference=f"Excel:{Path(self.file_path).name}",
                    )
                )

            # Remove valores importados anteriores
            self.db.query(AssumptionValue).filter(
                AssumptionValue.budget_version_id == self.budget_version_id,
                AssumptionValue.source_type == "imported",
            ).delete()

            self.db.add_all(values_to_insert)
            self.imported_count = len(values_to_insert)

            job.status = ImportStatus.completed if not self.errors else ImportStatus.partial
            job.summary = {
                "imported": self.imported_count,
                "errors": self.errors[:20],
                "sheets_processed": len(wb.sheetnames),
            }
        except Exception as e:
            job.status = "failed"
            job.error_detail = str(e)
            self.errors.append(str(e))
        finally:
            job.finished_at = datetime.now(timezone.utc)
            self.db.commit()

        return {
            "job_id": job_id,
            "status": job.status,
            "imported": self.imported_count,
            "errors": self.errors[:20],
        }

    def _process_sheet(self, ws, sheet_name: str) -> dict:
        """
        Processa uma aba e retorna {(code, period_or_None): value}.
        Estratégia: linha 1 = cabeçalhos (períodos), coluna A = labels.
        """
        results = {}

        # Lê cabeçalhos da linha 1
        headers = {}
        for cell in ws[1]:
            period = _extract_period_from_header(cell.value)
            if period:
                headers[cell.column] = period

        if not headers:
            # Sem períodos detectados — tenta ler como pares label:valor estático
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
                label_cell = row[0]
                val_cell = row[1] if len(row) > 1 else None

                if label_cell.value and val_cell and isinstance(val_cell.value, (int, float)):
                    code = _find_assumption_code(sheet_name, str(label_cell.value))
                    if code:
                        results[(code, None)] = val_cell.value
            return results

        # Com períodos: linha 1 = datas, coluna A = labels, resto = valores
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            label_cell = row[0]
            if not label_cell.value:
                continue

            code = _find_assumption_code(sheet_name, str(label_cell.value))
            if not code:
                continue

            for cell in row[1:]:
                if cell.column not in headers:
                    continue
                if isinstance(cell.value, (int, float)):
                    period = headers[cell.column]
                    results[(code, period)] = cell.value

        return results


def import_from_path(
    file_path: str,
    budget_version_id: str,
    db: "Session",
) -> dict:
    importer = ExcelImporter(file_path, budget_version_id, db)
    return importer.run()
